import openpyxl

# when you likely to create empty workbook
# wb = openpyxl.workbook()

# for loading existing workbook
wb = openpyxl.load_workbook("transactions.xlsx")

# gives you a sheetname
allsheet = wb.sheetnames

# worksheet set to sheet1 and its case senstive
sheet = wb["Sheet1"]

# creating new sheet before sheet 1 where 0 is indexing
"""wb.create_sheet("Sheet2",0)"""

# remvoing sheet
"""wb.remove_sheet(sheetname)"""

# getting value of an cell
cell = sheet["a1"]
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

"""alternate approach for sheet["a1"]"""
y = sheet.cell(row=1, column=1)
z = sheet.max_row
z1 = sheet.max_column

# using for loop

# here row is not 0 its start from 1 and for including last row we need to increase by 1

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


# get all column from cell a
column = sheet["a"]
for value in column:
    print(value.value)

# getting all column from given range
column = sheet["a:c"]
# printing column we get tuple in tuple
for row_tuple in column:
    for cell in row_tuple:
        print(cell.value)


# similarly for row
row = sheet[1]
# return all value of row 1 if used loop

# used to add row in the end of sheet
sheet.append([1, 2, 3])
"""
inserting row at given index
inserting col at given index
deleting row/col 
and many many more 
"""
# saving our workbook into new file named given below
wb.save("transaction2.xlsx")
